"""Excel (.xlsx) 存储 —— 需要 openpyxl，可选依赖。

  pip install openpyxl

每个 table 一个 sheet。一个 workbook 文件 mitm_rpc.xlsx。
"""
from __future__ import annotations

import json
import time
from pathlib import Path
from typing import Any

from ._base import Storage, register


@register("excel")
class ExcelStorage(Storage):
    def __init__(self, path: str | Path = "./data/mitm_rpc.xlsx"):
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            raise RuntimeError(
                "ExcelStorage 需要 openpyxl，安装: pip install openpyxl"
            )
        self.path = Path(path)
        self.path.parent.mkdir(parents=True, exist_ok=True)

    def _load(self):
        from openpyxl import Workbook, load_workbook
        if self.path.exists():
            return load_workbook(self.path)
        return Workbook()

    def _flatten(self, v: Any) -> Any:
        if isinstance(v, (dict, list)):
            return json.dumps(v, ensure_ascii=False)
        return v

    def write(self, table: str, record: dict[str, Any]) -> None:
        from openpyxl import Workbook
        record = {"_ts": int(time.time() * 1000), **record}
        wb = self._load()
        if table in wb.sheetnames:
            ws = wb[table]
            # 头在第 1 行
            headers = [c.value for c in ws[1]] if ws.max_row else []
        else:
            # 干净的 default sheet 删掉
            if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1 and wb["Sheet"].max_row <= 1:
                del wb["Sheet"]
            ws = wb.create_sheet(table)
            headers = []

        # 扩展 headers
        for k in record:
            if k not in headers:
                headers.append(k)
        # 写头
        for i, h in enumerate(headers, 1):
            ws.cell(1, i, h)
        # 写一行
        new_row = ws.max_row + 1
        for i, h in enumerate(headers, 1):
            ws.cell(new_row, i, self._flatten(record.get(h)))

        wb.save(self.path)

    def read(self, table: str, limit: int = 100, **filters) -> list[dict[str, Any]]:
        if not self.path.exists():
            return []
        wb = self._load()
        if table not in wb.sheetnames:
            return []
        ws = wb[table]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = rows[0]
        out: list[dict[str, Any]] = []
        for r in rows[1:]:
            rec = {h: v for h, v in zip(headers, r) if h is not None}
            if filters and not all(rec.get(k) == v for k, v in filters.items()):
                continue
            out.append(rec)
        return out[-limit:]
