"""FastAPI 主服务：RPC bus + Web UI + 抓包/健康/操作 API。

启动:
    uvicorn core.server:app --host 127.0.0.1 --port 9999

或者用 scripts/start.ps1 一键起所有东西。
"""
from __future__ import annotations

import asyncio
import importlib
import json
import pkgutil
import sys
from pathlib import Path
from typing import Any

from fastapi import FastAPI, HTTPException, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, RedirectResponse

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from core.rpc import RpcBus, BrowserSession  # noqa: E402
from sites._base import SitePlugin  # noqa: E402
from storage import get_storage, list_backends, STORAGE_REGISTRY  # noqa: E402
import pipelines  # noqa: E402

# 加载用户的 pipelines/ 处理器
pipelines.load_all()

# ─────────────────── 全局对象 ───────────────────

app = FastAPI(title="mitm-rpc")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"], allow_methods=["*"], allow_headers=["*"], allow_credentials=False,
)

bus = RpcBus()

# 存储后端：从 data/storage.config.json 读，没有就用默认 JSONL
STORAGE_CONFIG_FILE = ROOT / "data" / "storage.config.json"

def _load_storage_config() -> dict:
    if STORAGE_CONFIG_FILE.exists():
        try:
            return json.loads(STORAGE_CONFIG_FILE.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {"backend": "jsonl", "options": {"root": str(ROOT / "data" / "storage")}}

def _save_storage_config(cfg: dict) -> None:
    STORAGE_CONFIG_FILE.parent.mkdir(parents=True, exist_ok=True)
    STORAGE_CONFIG_FILE.write_text(json.dumps(cfg, ensure_ascii=False, indent=2),
                                   encoding="utf-8")

_storage_config = _load_storage_config()
try:
    storage = get_storage(_storage_config["backend"], **_storage_config.get("options", {}))
except Exception as e:
    print(f"[storage] '{_storage_config['backend']}' 启动失败 ({e})，回退到 jsonl")
    _storage_config = {"backend": "jsonl", "options": {"root": str(ROOT / "data" / "storage")}}
    storage = get_storage("jsonl", root=str(ROOT / "data" / "storage"))

# 加载站点插件
PLUGINS: dict[str, SitePlugin] = {}
def _load_plugins() -> None:
    sites_pkg = importlib.import_module("sites")
    for _f, name, is_pkg in pkgutil.iter_modules(sites_pkg.__path__):
        if not is_pkg or name.startswith("_"):
            continue
        try:
            mod = importlib.import_module(f"sites.{name}")
        except Exception as e:
            print(f"[plugin] load fail {name}: {e}")
            continue
        plugin = getattr(mod, "PLUGIN", None)
        if plugin is None:
            for attr in dir(mod):
                obj = getattr(mod, attr)
                if isinstance(obj, type) and issubclass(obj, SitePlugin) and obj is not SitePlugin:
                    plugin = obj()
                    break
        if plugin is not None:
            PLUGINS[plugin.name] = plugin
            print(f"[plugin] loaded: {plugin.name}")
_load_plugins()

# ─────────────────── 抓包开关（与 mitm_addon 共享文件标志） ───────────────────

CAPTURE_DIR = ROOT / "data" / "captures"
CAPTURE_DIR.mkdir(parents=True, exist_ok=True)
CAPTURE_FLAG = CAPTURE_DIR / "_enabled"  # 文件存在 = 开
CAPTURE_FILE = CAPTURE_DIR / "live.jsonl"

# ─────────────────── RPC 路由 ───────────────────

@app.post("/rpc/req")
async def rpc_req(req: Request):
    body = await req.json()
    op = body.pop("op", None)
    if not op:
        raise HTTPException(400, "missing op")
    timeout = body.pop("_timeout", 25.0)
    try:
        result = await bus.send(op, timeout=timeout, **body)
        return JSONResponse(result)
    except HTTPException:
        raise


@app.get("/rpc/poll")
async def rpc_poll():
    task = await bus.poll(wait=5.0)
    return JSONResponse(task or {})


@app.post("/rpc/result/{task_id}")
async def rpc_result(task_id: str, req: Request):
    body = await req.json()
    ok = bus.deliver(task_id, body)
    return {"ok": ok}


# ─────────────────── 站点 / 操作 API ───────────────────

@app.get("/api/sites")
async def list_sites():
    out = []
    for name, p in PLUGINS.items():
        out.append({
            "name": name,
            "domains": p.domains,
            "operations": list(p.operations().keys()),
            "patches": [{"name": pt.name, "notes": pt.notes} for pt in p.patches],
        })
    return out


@app.post("/api/sites/{site}/op/{op_name}")
async def run_op(site: str, op_name: str, req: Request):
    plugin = PLUGINS.get(site)
    if plugin is None:
        raise HTTPException(404, f"site {site} not found")
    ops = plugin.operations()
    op = ops.get(op_name)
    if op is None:
        raise HTTPException(404, f"op {op_name} not found")
    body = await req.json() if req.headers.get("content-length") else {}
    sess = BrowserSession(bus=bus)
    try:
        result = op(sess, **body) if not asyncio.iscoroutinefunction(op) else await op(sess, **body)
    except Exception as e:
        raise HTTPException(500, f"{type(e).__name__}: {e}")
    # 自动持久化 + pipeline (用户钩子可过滤/转换/转发)
    if isinstance(result, dict) and result.get("_persist"):
        table = result.pop("_persist")
        items = result.get("items") or [result]
        kept = 0
        for it in items:
            ev = pipelines.emit("record", table=table, record=it)
            if ev is None:  # 被处理器丢弃
                continue
            processed = ev.get("record", it)
            storage.write(table, processed)
            kept += 1
        result.setdefault("_persisted", kept)
    # 业务级事件
    pipelines.emit(f"{op_name}:after", result=result)
    return result


# ─────────────────── 健康检查 ───────────────────

@app.get("/api/health/{site}")
async def health(site: str):
    plugin = PLUGINS.get(site)
    if plugin is None:
        raise HTTPException(404, f"site {site} not found")
    sess = BrowserSession(bus=bus)

    async def fetch_main_js() -> str | None:
        # 通过浏览器去取 main.js（带真实 cookie），再 patch 检查
        for d in plugin.domains:
            if not d.startswith("static."):
                continue
            try:
                r = await sess.fetch(f"https://{d}/", method="GET")
                if r.get("ok") and r.get("status") == 200:
                    return r.get("body", "")
            except Exception:
                pass
        return None

    # 注：很多 plugin 的 health_check 用 sync requests 而不是 RPC，简化逻辑
    rep = plugin.health_check(lambda: None)
    return {
        "site": rep.site, "ok": rep.ok, "ts": rep.timestamp,
        "patches_ok": rep.patches_ok, "patches_missing": rep.patches_missing,
        "fix_prompt": rep.fix_prompt, "detail": rep.detail,
    }


# ─────────────────── 抓包 API ───────────────────

@app.get("/api/capture/status")
async def capture_status():
    enabled = CAPTURE_FLAG.exists()
    n = sum(1 for _ in CAPTURE_FILE.open(encoding="utf-8")) if CAPTURE_FILE.exists() else 0
    return {"enabled": enabled, "count": n, "file": str(CAPTURE_FILE)}


@app.post("/api/capture/toggle")
async def capture_toggle(req: Request):
    body = await req.json() if req.headers.get("content-length") else {}
    on = bool(body.get("on", not CAPTURE_FLAG.exists()))
    if on:
        CAPTURE_FLAG.touch()
    else:
        CAPTURE_FLAG.unlink(missing_ok=True)
    return {"enabled": on}


@app.get("/api/capture/list")
async def capture_list(limit: int = 100, q: str = ""):
    if not CAPTURE_FILE.exists():
        return []
    rows = []
    for line in CAPTURE_FILE.read_text(encoding="utf-8").splitlines():
        try:
            r = json.loads(line)
        except json.JSONDecodeError:
            continue
        if q and q not in r.get("url", "") and q not in r.get("req_body", "")[:200]:
            continue
        rows.append({k: r.get(k) for k in ("ts", "method", "url", "resp_status")})
    return rows[-limit:]


@app.get("/api/capture/{idx}")
async def capture_get(idx: int):
    rows = []
    if CAPTURE_FILE.exists():
        for line in CAPTURE_FILE.read_text(encoding="utf-8").splitlines():
            try:
                rows.append(json.loads(line))
            except json.JSONDecodeError:
                continue
    if idx < 0 or idx >= len(rows):
        raise HTTPException(404, "not found")
    return rows[idx]


@app.post("/api/capture/clear")
async def capture_clear():
    if CAPTURE_FILE.exists():
        CAPTURE_FILE.unlink()
    return {"ok": True}


# ─────────────────── Storage API ───────────────────

@app.get("/api/storage/backends")
async def storage_backends():
    """列出所有可用的存储后端 + 当前使用的。"""
    backends = []
    for name in list_backends():
        cls = STORAGE_REGISTRY[name]
        # 提取构造函数参数名作为可配置选项
        import inspect
        sig = inspect.signature(cls.__init__)
        params = []
        for pname, p in sig.parameters.items():
            if pname == "self":
                continue
            params.append({
                "name": pname,
                "default": p.default if p.default is not inspect.Parameter.empty else None,
                "type": p.annotation.__name__ if hasattr(p.annotation, "__name__") else "str",
            })
        backends.append({"name": name, "params": params})
    return {"current": _storage_config, "backends": backends}


@app.post("/api/storage/config")
async def storage_set_config(req: Request):
    global storage, _storage_config
    body = await req.json()
    backend = body.get("backend", "jsonl")
    options = body.get("options", {})
    try:
        new = get_storage(backend, **options)
    except Exception as e:
        raise HTTPException(400, f"切换失败: {e}")
    try:
        storage.close()
    except Exception:
        pass
    storage = new
    _storage_config = {"backend": backend, "options": options}
    _save_storage_config(_storage_config)
    return {"ok": True, "config": _storage_config}


@app.get("/api/storage/tables")
async def storage_tables():
    """尝试列出已有 tables（jsonl/csv 看文件，sqlite/mysql 查 schema）。"""
    backend = _storage_config["backend"]
    tables = []
    try:
        if backend == "jsonl":
            d = Path(_storage_config["options"].get("root", "./data/storage"))
            tables = [p.stem for p in d.glob("*.jsonl")] if d.exists() else []
        elif backend == "csv":
            d = Path(_storage_config["options"].get("root", "./data/csv"))
            tables = [p.stem for p in d.glob("*.csv")] if d.exists() else []
        elif backend == "sqlite":
            import sqlite3
            p = _storage_config["options"].get("path", "./data.sqlite")
            if Path(p).exists():
                con = sqlite3.connect(p)
                cur = con.execute("SELECT DISTINCT table_name FROM records")
                tables = [r[0] for r in cur.fetchall()]
                con.close()
        elif backend == "excel":
            try:
                from openpyxl import load_workbook
                p = _storage_config["options"].get("path", "./data/mitm_rpc.xlsx")
                if Path(p).exists():
                    tables = load_workbook(p, read_only=True).sheetnames
            except Exception:
                pass
        elif backend == "mysql":
            try:
                import pymysql
                cfg = dict(_storage_config["options"])
                with pymysql.connect(**cfg) as conn, conn.cursor() as cur:
                    cur.execute("SHOW TABLES")
                    tables = [r[0] for r in cur.fetchall()]
            except Exception:
                pass
    except Exception:
        pass
    return tables


@app.get("/api/storage/{table}")
async def storage_read(table: str, limit: int = 100):
    return storage.read(table, limit=limit)


# ─────────────────── Pipeline / Hooks ───────────────────

@app.get("/api/pipelines")
async def pipelines_list():
    """列出所有已注册的 hook (从 pipelines/ 目录加载来的)。"""
    return {"hooks": pipelines.list_hooks()}


@app.post("/api/pipelines/reload")
async def pipelines_reload():
    """重新扫描 pipelines/ 目录加载新 hook (改文件后调用)。"""
    pipelines._LOADED = False
    pipelines._HOOKS.clear()
    pipelines.load_all()
    return {"ok": True, "count": len(pipelines.list_hooks())}


# ─────────────────── 抓包内部入口 + SSE 流 ───────────────────

# 用 deque 同时供 SSE 推送 + HTTP 拉取
from collections import deque  # noqa: E402
_capture_queue: deque[dict] = deque(maxlen=1000)
_capture_listeners: list[asyncio.Queue] = []


@app.post("/api/internal/capture")
async def internal_capture(req: Request):
    """mitm addon 把每条捕获 POST 进来。会触发 pipeline + SSE 推送。"""
    rec = await req.json()
    # 跑 pipeline (可丢弃 / 改写 / 转发)
    ev = pipelines.emit("capture", record=rec)
    if ev is None:
        return {"ok": True, "dropped": True}
    rec = ev.get("record", rec) if isinstance(ev, dict) else rec
    _capture_queue.append(rec)
    # 推送给所有 SSE 订阅者
    for q in list(_capture_listeners):
        try:
            q.put_nowait(rec)
        except asyncio.QueueFull:
            pass
    # 持久化到文件 (供 /api/capture/list 拉)
    if CAPTURE_FLAG.exists():
        try:
            with CAPTURE_FILE.open("a", encoding="utf-8") as f:
                f.write(json.dumps(rec, ensure_ascii=False) + "\n")
        except Exception:
            pass
    return {"ok": True}


from fastapi.responses import StreamingResponse  # noqa: E402

@app.get("/api/capture/stream")
async def capture_stream():
    """SSE 实时推送新抓包记录。前端: new EventSource('/api/capture/stream')"""
    q: asyncio.Queue = asyncio.Queue(maxsize=200)
    _capture_listeners.append(q)

    async def gen():
        try:
            while True:
                rec = await q.get()
                yield f"data: {json.dumps(rec, ensure_ascii=False)}\n\n"
        finally:
            try:
                _capture_listeners.remove(q)
            except ValueError:
                pass

    return StreamingResponse(gen(), media_type="text/event-stream")


# ─────────────────── 统计 ───────────────────

@app.get("/api/stats")
async def stats():
    return {
        "rpc": bus.stats,
        "sites": list(PLUGINS.keys()),
        "capture_enabled": CAPTURE_FLAG.exists(),
    }


# ─────────────────── 首页 ───────────────────

@app.get("/")
async def index():
    """无 UI，返回服务状态 + 引导到 /docs (FastAPI 自动生成的 OpenAPI 页)"""
    return {
        "name": "mitm-rpc",
        "sites": list(PLUGINS.keys()),
        "storage": _storage_config,
        "rpc": bus.stats,
        "capture_enabled": CAPTURE_FLAG.exists(),
        "endpoints": {
            "api_docs": "/docs",
            "rpc": "/rpc/req",
            "capture_stream": "/api/capture/stream (SSE)",
            "pipelines": "/api/pipelines",
        },
        "cli": "python cli.py --help",
    }
